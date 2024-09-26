############################################################################
############################################################################
####    Explanation
############################################################################
############################################################################
###     2 Environments: Streamlit, Google Drive
###
###         Streamlit: 3 Containers
###             1) Upload Form
###             2) Review List
###             3) Review Comparison
###         
###         Google Drive: 3 Containers
###             1) Submission Metadata - Sheet
###             2) Submitted Forecast #s - Sheet
###             3) File Storage for all versions - Folder
###         
###         Processes
###         
###         




############################################################################
############################################################################
####    Initializing
############################################################################
############################################################################

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload
from tempfile import NamedTemporaryFile
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_interval
import re
from datetime import date
from datetime import datetime, timedelta
from io import BytesIO
import io
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
import os

pd.options.mode.chained_assignment = None  # default='warn'

ssid_subm = '10Od0nhz92hKVpuAwXFg45XGRZMz6x8vYf3cPvtxQe5E'
ssid_full = '1SsrJp5370HOfCURm64vOHIjZnMxmeC5deVZ6bLTEJE4'


creds = service_account.Credentials.from_service_account_file(
          'serviceacc.json',
          scopes=['https://www.googleapis.com/auth/drive.file',
                  'https://www.googleapis.com/auth/drive',
                  'https://www.googleapis.com/auth/drive.metadata',
                  'https://www.googleapis.com/auth/spreadsheets'
                  ]
          )  



############################################################################
############################################################################
####    Basic Variables / Formatting
############################################################################
############################################################################

def original_v_clean_message():
    themessage = """
    ---
    :gray['**Original**' refers to the file as it was uploaded, with no changes made.  
    '**Clean**' refers to a version of the file with *only* the numbers included and all formulas stripped (except for the summary, those are automatically added).  
    Please refer to the Budget files for any of the Service Lines or the initial email sent out for an example.]"""
    return themessage

def today_string():
    today = datetime.today() - timedelta(hours=5)
    current_date = today.strftime('%Y-%m-%d')
    return current_date

def today_string_file():
    today = datetime.today() - timedelta(hours=5)
    current_date = today.strftime('%m%d')
    return current_date
    
def number_naming_convention(num):
   nums = str(num)
   return_str = ''
   for x in range(4-len(nums)):
       return_str += '0'
   return return_str + nums




############################################################################
############################################################################
####    Foundational Functions (used in other Functions)
############################################################################
############################################################################

def GET_data_from_googlesheet(spreadsheetId, spreadsheetRange):   
    service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
    spreadsheetId = spreadsheetId
    spreadsheetRange = spreadsheetRange
    result = service.spreadsheets().values().get(
                                                spreadsheetId=spreadsheetId, 
                                                range=spreadsheetRange
                                                ).execute()
    df = pd.DataFrame(result['values'])
    df.columns = df.iloc[0]
    dfpiv = df[1:]
    return df, dfpiv
    
    
def APPEND_clean_version_detailed_data_to_master_table(spreadsheetId, spreadsheetRange, valueBody, inputOption='USER_ENTERED'):  
    service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
    spreadsheetId = spreadsheetId  # '1-zYgl-7ffj8cV2N80aICDHHKHfqyQX5rE3HXDcgSsfc'
    spreadsheetRange = spreadsheetRange  # 'CurrentFacilityValues!A:BQ'
    inputOption = 'USER_ENTERED'
    valueBody = valueBody
    result = service.spreadsheets().values().append(
                                              spreadsheetId=spreadsheetId, 
                                              range=spreadsheetRange,
                                              valueInputOption=inputOption, 
                                              body=valueBody
                                              ).execute()
    


def excel_storage_conversion(df):
    goog = df.values.tolist()
    return { 'values': goog }




############################################################################
############################################################################
####    Comparison Tool Functions
############################################################################
############################################################################

def from_SubmissionTitle_return_SL_FM_IN(submission_title):
   subm_df = GET_data_from_googlesheet(ssid_subm, 'All!A1:K')[0]
   try:
    filtered_list = subm_df[(subm_df['SubmissionTitle']==submission_title)].reset_index(drop=True)
    sl_found = filtered_list['ServiceLine'][0]
    fm_found = filtered_list['Version'][0]
    in_found = filtered_list['Iteration'][0]
   except:
    sl_found = ''
    fm_found = ''
    in_found = ''
   return sl_found, fm_found, in_found

def query_current_and_previous_version_ids(service_line, forecast_month, iteration_num):
   subm_df = GET_data_from_googlesheet(ssid_subm, 'All!A1:K')[0]
   # Current ID
   try:
        filtered_list1 = subm_df[(subm_df['ServiceLine']==service_line) & (subm_df['Version']==str(forecast_month)) & (subm_df['Iteration']==str(iteration_num))].reset_index(drop=True)
        current_id = filtered_list1['SubmissionID'][0]
   except:
        current_id = len(filtered_list1['SubmissionID']) #'fail'   

   # Previous ID
   try:
       filtered_list2 = subm_df[(subm_df['ServiceLine']==service_line) & (subm_df['Version']==forecast_month) & (subm_df['Iteration']==str(int(iteration_num)-1))].reset_index(drop=True)
       previous_id = filtered_list2['SubmissionID'][0]
   except:
       previous_id = '' 
   return current_id, previous_id    # filtered_list1, filtered_list2 #

def generate_list_within_forecast_month(service_line, forecast_month):
   subm_df = GET_data_from_googlesheet(ssid_subm, 'All!A1:K')[0]
   try:
    filtered_list = subm_df[(subm_df['ServiceLine']==service_line) & (subm_df['Version']==forecast_month)]
    versions_to_compare = list(filtered_list['SubmissionTitle'])
   except:
    versions_to_compare = []
   return versions_to_compare

def get_df_from_full_dataset_using_subid(subm_id, service_line):
   # try:
      full_df = GET_data_from_googlesheet(ssid_full, service_line+'!A1:P')[0]
      specified_df = full_df[(full_df['submission_id']==subm_id)]
      specified_df = specified_df.drop('submission_id', axis=1)
      return specified_df.reset_index(drop=True)
   # except:
    #   return pd.DataFrame()

def generate_df_changes(df1, df2, service_line, forecast_month):  #, abus_flag):
    try:
        diff = df1.compare(df2)
        diffT = diff.T
        col_mon = diff.columns.to_list()
        col_exa = diff.T.columns.to_list()
        
        # os.write(1, f"{x}\n".encode()) 
        # os.write(1, "watch\n".encode()) 
        
        if service_line == 'Mamm' and forecast_month in ('Budget','00+12','01+11','02+10','03+09','04+08','05+07','06+06'):
            exam_ref = ['Screening Mammography', 'Screening Breast US', 'Diagnostic Mamm', 
                    'Recall from Screening', 'Ductogram', 'Breast Ultrasound', 'Biopsy', 
                    'Stereotactic Biopsy', 'Breast MRI Biopsy', 'Breast MRI', 'DEXA', 
                    'Needle Loc', 'Seed Loc', 'Abscess Drainage', 'Sentinel Injection', 
                    'Cyst Aspiration']
        elif service_line == 'Mamm' and forecast_month in ('07+05','08+04','09+03','10+02','11+01','12+00','Strat Plan'):
            exam_ref = ['Screening Mammography', 'Screening Breast US', 
                        'ABUS Screening Breast US',
                         'Diagnostic Mamm', 
                    'Recall from Screening', 
                    'Recall from ABUS Screening Dx Mamm',
                    'Ductogram', 'Breast Ultrasound', 
                    'Recall from ABUS Screening US (ltd & cmp)',
                    'ABUS Dx Breast US (ltd & cmp)',
                    'Biopsy', 
                    'Stereotactic Biopsy', 'Breast MRI Biopsy', 'Breast MRI', 'DEXA', 
                    'Needle Loc', 'Seed Loc', 'Abscess Drainage', 'Sentinel Injection', 
                    'Cyst Aspiration']
        elif service_line == 'CIS':
            exam_ref = ['CT', 'Cal Sc CT', 'Cardiac CT', 'DX', 'MR', 'US', 'Fluoroscopy', 
                    'Screening Mamm', 'DEXA']
        else:
            exam_ref = ['New Patient Consults', '1st Veins', 'Additional Veins', 
                    'MD Sclerotherapy', 'Ultrasounds', 'Other']
 
        # flist_int = list(range(0,len))
        elist_int = list(range(0,len(col_exa)))
        mlist_int = list(range(0,len(col_mon),2))

        if service_line == 'Mamm' and forecast_month in ('Budget','00+12','01+11','02+10','03+09','04+08','05+07','06+06'):
            exam_type_num = 16
        elif service_line == 'Mamm' and forecast_month in ('07+05','08+04','09+03','10+02','11+01','12+00','Strat Plan'):
            exam_type_num = 20
        elif service_line == 'CIS':
            exam_type_num = 9
        else:
            exam_type_num = 6

        string_output = ''
        for row in elist_int:
            for col in mlist_int:
                from_value = str(diff.iloc[row,col+1]).replace(',','')
                to_value = str(diff.iloc[row,col]).replace(',','')
                if pd.isna(diff.iloc[row,col]) == False:
                    # debugval = diff.iloc[row,col]
                    # os.write(1, f"{debugval}\n".encode()) 
                    exam_row = col_exa[row] % exam_type_num
                    new_line = '*  ' + df1['FacilityName'][col_exa[row]] + '  ///  ' + exam_ref[exam_row] + '  (' + col_mon[col][0] + '):  from  ' + \
                        str(round(float(from_value))) + '  →  ' + str(round(float(to_value))) + '\n'
                    
                    string_output = string_output + new_line
    except:
       string_output = 'No comparison available.'
    return string_output




############################################################################
############################################################################
####    Form Submission Functions:  In order
############################################################################
############################################################################

def excel_reader_get_data(excel_file, facility_list, service_line, forecast_month):
    def load_workbook_range(range_string, ws):
        col_start, col_end = re.findall("[A-Z]+", range_string)
        data_rows = []
        for row in ws[range_string]:
            data_rows.append([cell.value for cell in row])
        return pd.DataFrame(data_rows, columns=get_column_interval(col_start, col_end))

    if service_line == 'Mamm' and forecast_month in ('Budget','00+12','01+11','02+10','03+09','04+08','05+07','06+06'):
        range_sl = 'A1:N17'
    elif service_line == 'Mamm' and forecast_month in ('07+05','08+04','09+03','10+02','11+01','12+00','Strat Plan'):
        range_sl = 'A1:N21'
    elif service_line == 'CIS':
        range_sl = 'A1:N10'
    else:
        range_sl = 'A1:N7'
    wb = openpyxl.load_workbook(excel_file, data_only=True, read_only=True)
    data = {}
    for facility in facility_list:
       ws = wb[facility]
       data[facility] = load_workbook_range(range_string=range_sl, ws=ws)  # df
    return data  # dict of dfs

def upload_file_to_drive(upload_file, file_name):
  service = build("drive", "v3", credentials=creds)
  try: 
    file_metadata = {"name": file_name, "parents":["113OEOtoIU3iF3mQ4EwlZZGtPgbZfePqc"]}
    with NamedTemporaryFile(dir='.', suffix='.xlsx') as f:
      f.write(upload_file.getbuffer())
      media = MediaFileUpload(f.name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") 
    file = (
        service.files()
        .create(supportsTeamDrives=True, body=file_metadata, media_body=media, fields="id")
        .execute()
    )
    print(f'File ID: {file.get("id")}')

  except HttpError as error:
    print(f"An error occurred: {error}")
    file = None

  fileid = file.get("id")
  permission1 = {
       'type': 'user',
       'role': 'writer',
       'emailAddress': 'charlotteradiologydatateam@gmail.com',  # Please set your email of Google account.
   }
  service.permissions().create(fileId=fileid, body=permission1).execute()
  permission2 = {
        'type': 'anyone',
        'role': 'writer',
    }
  service.permissions().create(fileId=fileid, body=permission2).execute()
  
  return fileid


def create_clean_copy(new_file_name, df_dict, facility_list, service_line, forecast_month):
    def write_df_to_openpyxl_excel_sheet(df, ws, header=False, index=False, startrow=1, startcol=2):
        """Write DataFrame df to openpyxl worksheet ws"""
        rows = dataframe_to_rows(df, header=header, index=index)
        for r_idx, row in enumerate(rows, startrow + 1):
            for c_idx, value in enumerate(row, startcol + 1):
                 ws.cell(row=r_idx, column=c_idx).value = value

    buffer = io.BytesIO()
    service = build("drive", "v3", credentials=creds)
    if service_line == 'Mamm' and forecast_month in ('Budget','00+12','01+11','02+10','03+09','04+08','05+07','06+06'):
        template_name = 'Mamm_Template.xlsx'
    elif service_line == 'Mamm' and forecast_month in ('07+05','08+04','09+03','10+02','11+01','12+00'):
        template_name = 'Mamm_Template_ABUS.xlsx'
    elif service_line == 'Mamm' and forecast_month in ('Strat Plan'):
        template_name = 'Mamm_Template_Strat.xlsx'
    elif service_line == 'CIS' and forecast_month in ('Strat Plan'):
        template_name = 'CIS_Template_Strat.xlsx'
    elif service_line == 'CIS':
        template_name = 'CIS_Template.xlsx'
    elif service_line == 'Vein' and forecast_month in ('Strat Plan'):
        template_name = 'Vein_Template_Strat.xlsx'
    else:
        template_name = 'Vein_Template.xlsx'
    try:
        wb = load_workbook(template_name)
        for facility in facility_list:
            df = df_dict[facility].iloc[1:,2:]   ##########
            ws = wb[facility] 
            write_df_to_openpyxl_excel_sheet(df, ws) 
        wb.save(buffer)
        buffervals = buffer.getvalue()
        
        file_metadata = {"name": new_file_name + ' (clean).xlsx', "parents":["113OEOtoIU3iF3mQ4EwlZZGtPgbZfePqc"]}
        with NamedTemporaryFile(dir='.', suffix='.xlsx') as f:
            f.write(buffervals)
            media = MediaFileUpload(f.name, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet") 
        file = (
            service.files()
            .create(supportsTeamDrives=True, body=file_metadata, media_body=media, fields="id")
            .execute()
        )
    except HttpError as error:
        print(f"An error occurred: {error}")
        file = None

    fileid = file.get("id")
    permission1 = {
        'type': 'user',
        'role': 'writer',
        'emailAddress': 'charlotteradiologydatateam@gmail.com',  # Please set your email of Google account.
    }
    service.permissions().create(fileId=fileid, body=permission1).execute()
    permission2 = {
            'type': 'anyone',
            'role': 'writer',
        }
    service.permissions().create(fileId=fileid, body=permission2).execute()
    
    return fileid

def final_combine_and_store_all_facilities(df_dict, facility_list, submission_id, serviceline, forecast_select, itnum):
    def reformat_add_df_context(df, facility, submission_id, forecast_select, itnum):
        df['Facility'] = facility
        df['submission_id'] = submission_id
        df['ForecastVersion'] = forecast_select
        df['IterationNum'] = itnum
        df=df[1:]
        return df
    for facility in facility_list:
      df = df_dict[facility]
      df = reformat_add_df_context(df, facility, submission_id, forecast_select, itnum)
      body = excel_storage_conversion(df)
      APPEND_clean_version_detailed_data_to_master_table(ssid_full, serviceline+'!A:R', body) ############## in progress

def add_submission_line(metadata):
    metadata_df = pd.DataFrame(metadata, index=[0])
    goog = excel_storage_conversion(metadata_df)
    APPEND_clean_version_detailed_data_to_master_table(ssid_subm, 'All!A:K', goog)
    
def get_iteration(service_line, forecast_month):
   subm_df = GET_data_from_googlesheet(ssid_subm, 'All!A1:K')[0]
   filtered_list = subm_df[(subm_df['ServiceLine']==service_line) & (subm_df['Version']==forecast_month)]
   return len(filtered_list)



############################################################################
############################################################################
####    File-Based Functions
############################################################################
############################################################################




############################################################################
############################################################################
####    No longer needed?
############################################################################
############################################################################

def DEPRECATE_stored_SET_data(spreadsheetId, spreadsheetRange, valueBody, inputOption='USER_ENTERED'):  
    service = build('sheets', 'v4', credentials=creds, cache_discovery=False)
    spreadsheetId = spreadsheetId  # '1-zYgl-7ffj8cV2N80aICDHHKHfqyQX5rE3HXDcgSsfc'
    spreadsheetRange = spreadsheetRange  # 'CurrentFacilityValues!A1:BQ321'
    inputOption = 'USER_ENTERED'
    valueBody = valueBody
    result = service.spreadsheets().values().update(
                                              spreadsheetId=spreadsheetId, 
                                              range=spreadsheetRange,
                                              valueInputOption=inputOption, 
                                              body=valueBody
                                              ).execute()
                                              
def DEPRECATE_convert_df(excel_file, facility_list, service_line):
    output = BytesIO()
    df_dict = excel_reader_get_data(excel_file, facility_list, service_line)
    writer = pd.ExcelWriter(output, 
                            engine='xlsxwriter', 
                            engine_kwargs={'options':{'strings_to_numbers':True, 'in_memory': True}})
    for i in range(len(facility_list)):
        df_dict[df_dict['FacilityName']==facility_list[i]].to_excel(writer,
                                                                 sheet_name=facility_list[i],
                                                                 index=False)
    writer.close()
    return output.getvalue()
